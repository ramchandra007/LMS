from django.shortcuts import render ,get_object_or_404 , redirect 
from . models import *

from datetime import datetime, timedelta, timezone
from django.http import FileResponse, HttpResponse,HttpResponseRedirect
from django.utils import timezone
from django.contrib import messages

def teacher_sidebar(request):      
    
    k2=leave.objects.filter(admin=request.user.id,read1=0) 
    leave_count = leave.objects.filter(admin=request.user.id,read1=0).count() 
    login=Teachers.objects.get(admin=request.user.id) 

    k4=compose_message.objects.filter(teachername=login)   
    loginteach=Teachers.objects.get(admin=request.user.id)   
    k=Teacher_Shifts.objects.filter(facult_name=loginteach) 
    k1=compose_message.objects.filter(teachername=login,read1=0)             

    k7=compose_message.objects.filter(teachername=login,read1=0)  
    message_count = compose_message.objects.filter(teachername=login,read1=0).count()  
    current_time = timezone.now()
    total_count=leave_count+message_count;                                                                                                                                          
  
    k5=teachermenu.objects.all()                                                                                                                                                                                                                                                                                                                                                                         
    return render(request,'teacher-template/teachersidebar.html',{'k':k,'k1':k1,'k5':k5,'k2':k2,'leave_count':leave_count,'k7':k7,'message_count':message_count,'total_count':total_count,'k4':k4,'current_time':current_time})




def teacher_applyforleave(request):
    k2=leave.objects.filter(admin=request.user.id)
    k5=teachermenu.objects.all()

    k1=leavestype.objects.all()
    adminid=CustomUser.objects.get(id=request.user.id)
    pending_leave_count = leave.objects.filter(is_status=0,admin=request.user.id).count()
    approved_leave_count = leave.objects.filter(is_status=1,admin=request.user.id).count()
    disapproved_leave_count = leave.objects.filter(is_status=2,admin=request.user.id).count()

    if request.method == "POST":
        Leave_Type=request.POST.getlist('Leave_Type')
        Reason=request.POST.get('Reason')
        from_date_str = request.POST.get('from_date')
        to_date_str = request.POST.get('to_date')

        from_date = datetime.strptime(from_date_str, '%Y-%m-%dT%H:%M')
        to_date = datetime.strptime(to_date_str, '%Y-%m-%dT%H:%M')

        days_difference = (to_date - from_date).days

        for Leave_Type, from_date_str, to_date_str in zip(Leave_Type, from_date_str,to_date_str):
            if Leave_Type and from_date_str and to_date_str:
                Leave_instance = leavestype.objects.get(id=Leave_Type)
        k3=leave(Leave_Type=Leave_instance,Reason=Reason,from_date=from_date,to_date=to_date,admin=adminid,user_type=2,days_difference=days_difference)
        k3.save()
    return render(request,'teacher-template/teacher_apply_leave.html',{'k5':k5,'k1':k1,'k2':k2,'pending_leave_count':pending_leave_count,'approved_leave_count':approved_leave_count,'disapproved_leave_count':disapproved_leave_count})

def mark_all_as_read1(request):
    leave.objects.filter(admin=request.user.id).update(read1=True)
    return HttpResponseRedirect('/teacher_sidebar') 

def teacher_mark_as_read(request, leave_id):
    leave_obj = get_object_or_404(leave, id=leave_id)
    leave_obj.read1 = True
    leave_obj.save()
    return HttpResponseRedirect('/teacher_sidebar')

from .models import teachermenu
def teacher_sidebar2(request):
    # login=Teachers.objects.get(admin=request.user.id)
    # k2=compose_message.objects.filter(teachername=login)
    # leave_count = compose_message.objects.count()

    k5=teachermenu.objects.filter(parent_category=None)
    k2=compose_message.objects.filter(teachername__admin=request.user.id)
    for message in k2:
        if message.MessageType == "0":
            message.ShortMessage = message.Message
        if message.MessageType == "1":
            message.ShortMessage = message.Message
        if message.MessageType == "2":
            message.ShortMessage = message.Message
        if message.MessageType == "3":
            message.ShortMessage = message.Message    




    return render(request,'teacher-template/teacher_sidebar2.html',{'k2':k2,'k5':k5})
def teacher_messages_mark_all_as_read(request): 
    login=Teachers.objects.get(admin=request.user.id) 

    compose_message.objects.filter(teachername=login).update(read1=True) 
    return HttpResponseRedirect('/teacher_sidebar')  
def teacher_messages_mark_as_read(request, compose_message_id):
    compose_message_obj = get_object_or_404(compose_message, id=compose_message_id)
    compose_message_obj.read1 = True
    compose_message_obj.save()                                                     
    return HttpResponseRedirect('/teacher_sidebar2')  





def dealing_subjects(request):
    k5=teachermenu.objects.filter(parent_category=None)
    k=Teachers.objects.get(admin=request.user.id)
    k0=Teacher_Class_sub.objects.filter(teacher=k)
    return render(request,'teacher-template/dealing_subjects.html',{'k0':k0,'k5':k5})


def Access_subjects(request):
    k5=teachermenu.objects.filter(parent_category=None)
    k=Teachers.objects.get(admin=request.user.id)
    k0=Teacher_Class_sub.objects.filter(teacher=k)
    return render(request,'teacher-template/Access_subjects.html',{'k0':k0,'k5':k5})


def Set_quiz(request,id):
    k2=Teacher_Class_sub.objects.get(id=id)
    k2.is_control=1
    k2.save()
    return HttpResponseRedirect('/dealing_subjects')

def Not_set_quiz(request,id):
    k2=Teacher_Class_sub.objects.get(id=id)
    k2.is_control=0
    k2.save()
    return HttpResponseRedirect('/dealing_subjects')



from django.contrib import messages
def quiz(request,id):

    form = MyForm()

    k0=Teacher_Class_sub.objects.get(id=id)
    if request.method == "POST":
        question=request.POST.get('question')
        choice1=request.POST.get('choice1')
        choice2=request.POST.get('choice2')
        choice3=request.POST.get('choice3')
        choice4=request.POST.get('choice4')
        is_correct=request.POST.get('is_correct')
        class_name=request.POST.get('class_name')
        subject=request.POST.get('subject')
        classes_instance=cls_name.objects.get(id=class_name)
        subject_instance=Subject.objects.get(id=subject)
        hk=quiz_questions(question=question,choice1=choice1,choice2=choice2,choice3=choice3,choice4=choice4,is_correct=is_correct,class_name=classes_instance,subject=subject_instance)
        hk.save()   
        messages.success(request,"Question inserted successfully")
    return render(request, 'teacher-template/quiz.html',{'k0':k0,'form':form})



def time(request,id):
    k0=Teacher_Class_sub.objects.get(id=id)
    if request.method=="POST":
        houres=request.POST.get('houres')
        minutes=request.POST.get('minutes')
        secondes=request.POST.get('secondes')
        K=set_timer(houres=houres,minutes=minutes,secondes=secondes,class_name=k0.class_name,subject=k0.subject)
        K.save()   
        messages.success(request,"Time Updated Successfully")
    return render(request,"teacher-template/time.html",{'k0':k0})



from django.shortcuts import render
from openpyxl import load_workbook



def import_from_excel(request,id):
    k0=Teacher_Class_sub.objects.get(id=id)
    m=Excel.objects.all()
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']
        class_name_id = request.POST.get('class_name')
        subject_id = request.POST.get('subject')

        wb = load_workbook(excel_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
      
            question,choice1,choice2,choice3,choice4,is_correct = row
            quiz_questions.objects.create(is_correct=is_correct,question=question, choice1=choice1,choice2=choice2,choice3=choice3,choice4=choice4,class_name_id=class_name_id,subject_id=subject_id)

        return HttpResponse("sucessfully upload")

    return render(request, 'teacher-template/excel.html',{'k0':k0,'m':m})


def insert_exel(request):
    if request.method=='POST':
        exel=request.FILES.get('exel')
        md=Excel(exel=exel)
        md.save()
        return HttpResponse("inserted.....")
    return render(request,"teacher-template/excel1.html")

def serve_pdf(request, document_id):
    document = get_object_or_404(Excel, pk=document_id)
    response = FileResponse(open(document.exel.path, 'rb'), content_type='application/excel')
    return response

def std_results(request):
    k5=teachermenu.objects.filter(parent_category=None)
    k=Teachers.objects.get(admin=request.user.id)
    k0=Teacher_Class_sub.objects.filter(teacher=k)
    return render(request,'teacher-template/std_results.html',{'k0':k0,'k5':k5})

